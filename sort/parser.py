import openpyxl
import re

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sort Benchmark Results"

def normalize(name: str):
    return " ".join(name.strip().split())

with open("results.txt", "r") as f:
    data = f.read()

file_pattern = re.compile(r'"([^"]+\.txt)"')
result_pattern = re.compile(r'^([^\n:]+):([\d.]+)', re.MULTILINE)

blocks = data.split('--------------------------')

all_algorithms = set()
for block in blocks:
    raw_results = result_pattern.findall(block)
    result_map = {normalize(name): time for name, time in raw_results}
    for name, _ in raw_results:
        all_algorithms.add(normalize(name))

algorithm_list = sorted(all_algorithms)

ws.cell(row=1, column=1).value = "File Name"
for col, algo in enumerate(algorithm_list, start=2):
    ws.cell(row=1, column=col).value = algo

row = 2
for block in blocks:
    if not block.strip():
        continue

    file_match = result_pattern.search(block)
    if not file_match:
        file_match=file_pattern.search(block)
        if not file_match:
            continue
        file_name = file_match.group(1)
        ws.cell(row=row, column=1).value = file_name
        continue



    raw_results = result_pattern.findall(block)
    result_map = {normalize(name): time for name, time in raw_results}
    for col, algo in enumerate(algorithm_list, start=2):
        normalized_algo = normalize(algo)
        time = result_map.get(normalized_algo)

        if time is not None:
            try:
                ws.cell(row=row, column=col).value = float(time)
            except ValueError as e:

    row += 1

wb.save("formatted_results.xlsx")
print("Excel file created successfully: formatted_results.xlsx")
